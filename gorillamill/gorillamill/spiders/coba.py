# import csv
# import os
# import scrapy
# import pandas as pd
# from openpyxl import Workbook, load_workbook

# from scrapy.selector import Selector
# from w3lib.html import remove_tags

# class GorillaMillSpider(scrapy.Spider):
#     name = "datacoba"
#     allowed_domains = ["gorillamill.com"]

#     def start_requests(self):
#         with open("Gorilla-Mill-Products.csv", 'r') as file:
#             reader = csv.reader(file)
#             next(reader)  # skip header row
#             for row in reader:
#                 try:
#                     product_id = row[0]
#                     url = f"https://gorillamill.com/products/product/{product_id}"
#                     yield scrapy.Request(url, self.parse, meta={"product_id": product_id})
#                 except:
#                     product_id = 'Id not working'

#     def parse(self, response):
#         itemlist = {}
#         title = " ".join(response.css('div#specs h2::text, h2::text, h1::text').getall())
#         title = title.strip().replace('\n', '').replace("\r", "").replace("\t", "") if title else ''
#         image = response.css('div#specs img.tool::attr(src)').get()
#         catalog_number = response.xpath('//div/strong[1]/following-sibling::text()').get()
#         # catalog_number = response.xpath('//div/strong[1]/following-sibling::text()').get()
#         checkstock = response.css('a[data-stock]::attr(data-stock)').get()
#         sel = Selector(text=response.text)

#         # Remove <a> tags
#         for a in sel.xpath('//div[@class="uk-width-1-1 uk-width-medium-4-10"]/a'):
#             a.extract()

#         # Get Desc text and remove line breaks and whitespace
#         desc = sel.xpath('//div[@class="uk-width-1-1 uk-width-medium-4-10"]/strong[contains(text(), "Desc:")]/following-sibling::text()[normalize-space()]')
#         desc = ''.join([d.strip().replace('\n', '').replace(' ', '') for d in desc.getall() if d.strip()])

#         itemlist = {
#             "title": title,
#             "image": image,
#             "catalog_number": catalog_number,
#             "Check Stock": checkstock,
#             'desc': desc
#         }
#         print(itemlist)

#         # Load workbook if exist, or create a new one
#         if os.path.exists('testfile.xlsx'):
#             book = load_workbook('testfile.xlsx')
#         else:
#             book = Workbook()

#         # Get sheet and append data
#         sheet_name = "Gorilla Mill Products"
#         if sheet_name not in book.sheetnames:
#             sheet = book.active
#             sheet.title = sheet_name
#             # add header row
#             header = ["Product ID", "catalog_number", "title", "image", "Check Stock", "desc"]
#             sheet.append(header)
#         else:
#             sheet = book[sheet_name]

#         # add data row
#         product_id = response.meta.get('product_id')
#         row = [product_id, itemlist['title'], itemlist['image'], itemlist['catalog_number'], itemlist['Check Stock'], itemlist['desc']]
#         sheet.append(row)

#         # Save the workbook
#         book.save('testfile.xlsx')



import csv
import os
import scrapy
import pandas as pd
from openpyxl import Workbook, load_workbook
from scrapy.selector import Selector
from w3lib.html import remove_tags

class GorillaMillSpider(scrapy.Spider):
    name = "datacoba"
    allowed_domains = ["gorillamill.com"]

    def start_requests(self):
        with open("Gorilla-Mill-Products.csv", 'r') as file:
            reader = csv.reader(file)
            next(reader)  # skip header row
            for row in reader:
                try:
                    product_id = row[0]
                    url = f"https://gorillamill.com/products/product/{product_id}"
                    yield scrapy.Request(url, self.parse, meta={"product_id": product_id})
                except:
                    product_id = 'Id not working'

    def parse(self, response):
        itemlist = {}
        title = " ".join(response.css('div#specs h2::text, h2::text, h1::text').getall())
        title = title.strip().replace('\n', '').replace("\r", "").replace("\t", "") if title else ''
        image = response.css('div#specs img.tool::attr(src)').get()
        catalog_number = response.xpath('//div/strong[1]/following-sibling::text()').get()
        checkstock = response.css('a[data-stock]::attr(data-stock)').get()
        desc = response.xpath('normalize-space(//div[@class="uk-width-1-1 uk-width-medium-4-10"])')
        desc = desc.get().replace('Check Stock >>', '').replace('Desc:', '').replace(',', '| ')  if desc else ''

        itemlist = {
            "title": title,
            "image": image,
            "catalog_number": catalog_number,
            "Check Stock": checkstock,
            'desc': desc
        }
        print(itemlist)

        # Load workbook if exist, or create a new one
        if os.path.exists('testfile2.xlsx'):
            book = load_workbook('testfile2.xlsx')
        else:
            book = Workbook()

        # Get sheet and append data
        sheet_name = "Gorilla Mill Products"
        if sheet_name not in book.sheetnames:
            sheet = book.active
            sheet.title = sheet_name
            # add header row
            header = ["Product ID", "catalog_number", "title", "image", "Check Stock", "desc"]
            sheet.append(header)
        else:
            sheet = book[sheet_name]

        # add data row
        product_id = response.meta.get('product_id')
        row = [product_id, itemlist['catalog_number'], itemlist['title'], itemlist['image'], itemlist['Check Stock'], itemlist['desc']]
        sheet.append(row)

        # Save the workbook
        try:
            book.save('testfile2.xlsx')
        except ValueError:
            for idx, val in enumerate(row):
                row[idx] = str(val)
            sheet.append(row)
            book.save('testfile2.xlsx')
