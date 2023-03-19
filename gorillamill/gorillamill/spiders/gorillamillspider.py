# import csv
# import scrapy
# import pandas as pd

# class GorillaMillSpider(scrapy.Spider):
#     name = "gorillamill"
#     allowed_domains = ["gorillamill.com"]
    

#     def start_requests(self):
#         with open("Gorilla-Mill-Products.csv", 'r') as file:
#             reader = csv.reader(file)
#             for row in reader:
#                 try:
#                     product_id = row[0]
#                     url = f"https://gorillamill.com/products/product/{product_id}"
#                 except:
#                     product_id = 'Id not working'
#                 yield scrapy.Request(url, self.parse)

#     def parse(self, response):
        
#         item = {}
       
# #         item['id'] = response.css('a[data-product]::attr(data-product)').get()
# #         item['title'] = " ".join(response.css('div#specs h2::text, h2::text, h1::text').getall())
# #         item['data stuck'] = response.css('a[data-stock]::attr(data-stock)').get()
# #         item['url'] = response.url
# #         print(item)
# #         yield item



# # import csv
# # import os
# # import scrapy
# # import pandas as pd

# # class GorillaMillSpider(scrapy.Spider):
# #     name = "gorillamill"
# #     allowed_domains = ["gorillamill.com"]

# #     def start_requests(self):
# #         with open("Gorilla-Mill-Products.csv", 'r') as file:
# #             reader = csv.reader(file)
# #             for row in reader:
# #                 try:
# #                     product_id = row[0]
# #                     url = f"https://gorillamill.com/products/product/{product_id}"
# #                 except:
# #                     product_id = 'Id not working'
# #                 yield scrapy.Request(url, self.parse)

# #     def parse(self, response):
# #         item = {}
# #         item['id'] = response.css('a[data-product]::attr(data-product)').get()
# #         item['title'] = " ".join(response.css('div#specs h2::text, h2::text, h1::text').getall())
# #         item['data stuck'] = response.css('a[data-stock]::attr(data-stock)').get()
# #         item['url'] = response.url
# #         yield item


# import csv
# import os
# import scrapy
# import pandas as pd
# from openpyxl import Workbook


# class GorillaMillSpider(scrapy.Spider):
#     name = "gorillamill"
#     allowed_domains = ["gorillamill.com"]

#     def start_requests(self):
#         with open("Gorilla-Mill-Products.csv", 'r') as file:
#             reader = csv.reader(file)
#             for row in reader:
#                 try:
#                     product_id = row[0]
#                     url = f"https://gorillamill.com/products/product/{product_id}"
#                 except:
#                     product_id = 'Id not working'
#                 yield scrapy.Request(url, self.parse)

#     def parse(self, response):
#         itemlist = []
#         title = " ".join(response.css('div#specs h2::text, h2::text, h1::text').getall())
#         image  =  response.css('div#specs img.tool::attr(src)').get()
#         catalog_number = response.xpath('//div/strong[1]/following-sibling::text()').get()
#         checkstock = response.css('a[data-stock]::attr(data-stock)').get()
        
#         itemlist = {
#             "title": title,
#             "image": image,
#             "catalog_number": catalog_number,
#             "Check Stock": checkstock
#         }



# import csv
# import os
# import scrapy
# import pandas as pd
# from openpyxl import Workbook


# class GorillaMillSpider(scrapy.Spider):
#     name = "gorillamill"
#     allowed_domains = ["gorillamill.com"]

#     def start_requests(self):
#         with open("Gorilla-Mill-Products.csv", 'r') as file:
#             reader = csv.reader(file)
#             for row in reader:
#                 try:
#                     product_id = row[0]
#                     url = f"https://gorillamill.com/products/product/{product_id}"
#                 except:
#                     product_id = 'Id not working'
#                 yield scrapy.Request(url, self.parse)

#     def parse(self, response):
#         itemlist = []
#         title = " ".join(response.css('div#specs h2::text, h2::text, h1::text').getall())
#         image = response.css('div#specs img.tool::attr(src)').get()
#         catalog_number = response.xpath('//div/strong[1]/following-sibling::text()').get()
#         checkstock = response.css('a[data-stock]::attr(data-stock)').get()

#         return {
#             'product_id': response.meta.get('product_id', ''),
#             "title": title,
#             "image": image,
#             "catalog_number": catalog_number,
#             "Check Stock": checkstock
#         }
        # print(item)
        # itemlist.append(item)

        # df = pd.DataFrame(itemlist, columns=['title', 'image', 'catalog_number', 'Check Stock'])
        # df.to_excel('list.xlsx', index=False)


# import csv
# import os
# import scrapy
# import pandas as pd
# from openpyxl import Workbook


# class GorillaMillSpider(scrapy.Spider):
#     name = "gorillamill"
#     allowed_domains = ["gorillamill.com"]

#     def start_requests(self):
#         with open("Gorilla-Mill-Products.csv", 'r') as file:
#             reader = csv.reader(file)
#             for row in reader:
#                 try:
#                     product_id = row[0]
#                     url = f"https://gorillamill.com/products/product/{product_id}"
#                 except:
#                     product_id = 'Id not working'
#                 yield scrapy.Request(url, self.parse, meta={'product_id': product_id})

#     def parse(self, response):
        
#         title = " ".join(response.css('div#specs h2::text, h2::text, h1::text').get()),
#         image  =  response.css('div#specs img.tool::attr(src)').get(),
#         catalog_number = response.xpath('//div/strong[1]/following-sibling::text()').get(),
#         checkstock = response.css('a[data-stock]::attr(data-stock)').get()
        
#         yield {
#             'product_id': response.meta.get('product_id', ''),
#             'title': title,
#             "image": image,
#             "catalog_number": catalog_number,
#             "Check Stock": checkstock
#         }
    
    # def closed(self, reason):
    #     items = list(self.crawler.stats.get('item_scraped_count', {}).values())
    #     if len(items) > 0:
    #         df = pd.DataFrame(items)
    #         df.to_excel('Gorilla-Mill-Products--0.xlsx', index=False)



# import csv
# import os
# import scrapy
# import pandas as pd
# # from openpyxl import Workbook
# from openpyxl import Workbook, load_workbook

# class GorillaMillSpider(scrapy.Spider):
#     name = "gorillamill"
#     allowed_domains = ["gorillamill.com"]

#     def start_requests(self):
#         with open("Gorilla-Mill-Products.csv", 'r') as file:
#             reader = csv.reader(file)
#             for row in reader:
#                 try:
#                     product_id = row[0]
#                     url = f"https://gorillamill.com/products/product/{product_id}"
#                 except:
#                     product_id = 'Id not working'
#                 yield scrapy.Request(url, self.parse)

#     def parse(self, response):
#         itemlist = []
#         title = " ".join(response.css('div#specs h2::text, h1::text').getall())
#         image  =  response.css('div#specs img.tool::attr(src)').get()
#         catalog_number = response.xpath('//div/strong[1]/following-sibling::text()').get()
#         checkstock = response.css('a[data-stock]::attr(data-stock)').get()
        
#         itemlist = {
#             "title": title,
#             "image": image,
#             "catalog_number": catalog_number,
#             "Check Stock": checkstock
#         }

#         # save data to Excel
#         df = pd.DataFrame([itemlist])
#         if os.path.exists('list.xlsx'):
#             book = load_workbook('list.xlsx')
#             writer = pd.ExcelWriter('list.xlsx', engine='openpyxl')
#             writer.book = book
#             writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
#         else:
#             writer = pd.ExcelWriter('list.xlsx', engine='openpyxl')
#             df.to_excel(writer, sheet_name='Sheet1', index=False)

#         df.to_excel(writer, sheet_name='Sheet1', index=False)
#         writer.save()


# import csv
# import os
# import scrapy
# import pandas as pd
# # from openpyxl import Workbook
# from openpyxl import Workbook, load_workbook

# from openpyxl import Workbook
# from openpyxl import load_workbook

# class GorillaMillSpider(scrapy.Spider):
#     name = "gorillamill"
#     allowed_domains = ["gorillamill.com"]

#     def start_requests(self):
#         with open("Gorilla-Mill-Products.csv", 'r') as file:
#             reader = csv.reader(file)
#             for row in reader:
#                 try:
#                     product_id = row[0]
#                     url = f"https://gorillamill.com/products/product/{product_id}"
#                 except:
#                     product_id = 'Id not working'
#                 yield scrapy.Request(url, self.parse)

#     def parse(self, response):
#         itemlist = []
#         title = " ".join(response.css('div#specs h2::text, h2::text, h1::text').getall())
#         image  =  response.css('div#specs img.tool::attr(src)').get()
#         catalog_number = response.xpath('//div/strong[1]/following-sibling::text()').get()
#         checkstock = response.css('a[data-stock]::attr(data-stock)').get()
        
#         itemlist = {
#             "title": title,
#             "image": image,
#             "catalog_number": catalog_number,
#             "Check Stock": checkstock
#         }
#         print(itemlist)

#         # Load workbook if exist, or create a new one
#         if os.path.exists('list.xlsx'):
#             book = load_workbook('list.xlsx')
#         else:
#             book = Workbook()

#         # Get sheet and append data
#         sheet_name = "Gorilla Mill Products"
#         if sheet_name not in book.sheetnames:
#             sheet = book.create_sheet(sheet_name)
#             # add header row
#             header = ["title", "image", "catalog_number", "Check Stock"]
#             sheet.append(header)
#         else:
#             sheet = book[sheet_name]

#         # add data row
#         row = [itemlist['title'], itemlist['image'], itemlist['catalog_number'], itemlist['Check Stock']]
#         sheet.append(row)

#         # Save the workbook
#         book.save('list.xlsx')


# import csv
# import os
# import scrapy
# import pandas as pd
# from openpyxl import Workbook, load_workbook

# class GorillaMillSpider(scrapy.Spider):
#     name = "gorillamill"
#     allowed_domains = ["gorillamill.com"]

#     def start_requests(self):
#         with open("Gorilla-Mill-Products.csv", 'r') as file:
#             reader = csv.reader(file)
#             for row in reader:
#                 try:
#                     product_id = row[0]
#                     url = f"https://gorillamill.com/products/product/{product_id}"
#                 except:
#                     product_id = 'Id not working'
#                 yield scrapy.Request(url, self.parse)

#     def parse(self, response):
#         itemlist = []
#         title = " ".join(response.css('div#specs h2::text, h2::text, h1::text').getall())
#         title = title.strip() if title else ''
#         image  =  response.css('div#specs img.tool::attr(src)').get()
#         catalog_number = response.xpath('//div/strong[1]/following-sibling::text()').get()
#         checkstock = response.css('a[data-stock]::attr(data-stock)').get()
        
#         itemlist = {
#             "title": title,
#             "image": image,
#             "catalog_number": catalog_number,
#             "Check Stock": checkstock
#         }
#         print(itemlist)

#         # Load workbook if exist, or create a new one
#         if os.path.exists('list.xlsx'):
#             book = load_workbook('list.xlsx')
#         else:
#             book = Workbook()

#         # Get sheet and append data
#         sheet_name = "Gorilla Mill Products"
#         if sheet_name not in book.sheetnames:
#             sheet = book.create_sheet(sheet_name)
#             # add header row
#             header = ["title", "image", "catalog_number", "Check Stock"]
#             sheet.append(header)
#         else:
#             sheet = book[sheet_name]

#         # add data row
#         row = [itemlist['title'], itemlist['image'], itemlist['catalog_number'], itemlist['Check Stock']]
#         sheet.append(row)

#         # Save the workbook
#         book.save('list.xlsx')



# import csv
# import os
# import scrapy
# import pandas as pd
# from openpyxl import Workbook, load_workbook

# class GorillaMillSpider(scrapy.Spider):
#     name = "gorillamill"
#     allowed_domains = ["gorillamill.com"]

#     def start_requests(self):
#         with open("Gorilla-Mill-Products.csv", 'r') as file:
#             reader = csv.reader(file)
#             for row in reader:
#                 try:
#                     product_id = row[0]
#                     url = f"https://gorillamill.com/products/product/{product_id}"
#                 except:
#                     product_id = 'Id not working'
#                 yield scrapy.Request(url, self.parse)

#     def parse(self, response):
#         itemlist = []
#         title = " ".join(response.css('div#specs h2::text, h2::text, h1::text').getall())
#         title = title.strip() if title else ''
#         image  =  response.css('div#specs img.tool::attr(src)').get()
#         catalog_number = response.xpath('//div/strong[1]/following-sibling::text()').get()
#         checkstock = response.css('a[data-stock]::attr(data-stock)').get()
        
#         itemlist = {
#             "title": title,
#             "image": image,
#             "catalog_number": catalog_number,
#             "Check Stock": checkstock
#         }
#         print(itemlist)

#         # Load workbook if exist, or create a new one
#         if os.path.exists('datalist.xlsx'):
#             book = load_workbook('datalist.xlsx')
#         else:
#             book = Workbook()

#         # Get sheet and append data
#         sheet_name = "Gorilla Mill Products"
#         if sheet_name not in book.sheetnames:
#             sheet = book.active
#             sheet.title = sheet_name
#             # add header row
#             header = ["title", "image", "catalog_number", "Check Stock"]
#             sheet.append(header)
#         else:
#             sheet = book[sheet_name]

#         # add data row
#         row = [itemlist['title'], itemlist['image'], itemlist['catalog_number'], itemlist['Check Stock']]
#         sheet.append(row)

#         # Save the workbook
#         book.save('datalist.xlsx')



import csv
import os
import scrapy
import pandas as pd
from openpyxl import Workbook, load_workbook

from scrapy.selector import Selector
from w3lib.html import remove_tags

class GorillaMillSpider(scrapy.Spider):
    name = "gorillamill"
    allowed_domains = ["gorillamill.com"]

    def start_requests(self):
        with open("Gorilla-Mill-Products.csv", 'r') as file:
            reader = csv.reader(file)
            next(reader)  # skip header row
            for row in reader:
                try:
                    product_id = row[0]
                    url = f"https://gorillamill.com/products/product/{product_id}"
                    yield scrapy.Request(url, self.parse, meta={"product_id": product_id})
                except:
                    product_id = 'Id not working'

    def parse(self, response):
        itemlist = []
        title = " ".join(response.css('div#specs h2::text, h2::text, h1::text').getall())
        title = title.strip() if title else ''
        image  =  response.css('div#specs img.tool::attr(src)').get()
        catalog_number = response.xpath('//div/strong[1]/following-sibling::text()').get()
        # catalog_number = response.xpath('//div/strong[1]/following-sibling::text()').get()
        checkstock = response.css('a[data-stock]::attr(data-stock)').get()
        sel = Selector(text=response.text)
        for a in sel.xpath('//div[@class="uk-width-1-1 uk-width-medium-4-10"]/a'):
            a.extract()

        # mengambil teks
        desc = sel.xpath('//div[@class="uk-width-1-1 uk-width-medium-4-10"]/strong[contains(text(), "Desc:")]/following-sibling::text()[normalize-space()]').getall()
        desc = " ".join(desc).replace("\n", "").replace("\r", "").replace("\t", "").strip()
        
        itemlist = {
            "title": title,
            "image": image,
            "catalog_number": catalog_number,
            "Check Stock": checkstock,
            'desc': desc
        }
        print(itemlist)

        # Load workbook if exist, or create a new one
        if os.path.exists('datalist.xlsx'):
            book = load_workbook('datalist.xlsx')
        else:
            book = Workbook()

        # Get sheet and append data
        sheet_name = "Gorilla Mill Products"
        if sheet_name not in book.sheetnames:
            sheet = book.active
            sheet.title = sheet_name
            # add header row
            header = ["Product ID", "catalog_number", "title", "image",  "Check Stock", "desc"]
            sheet.append(header)
        else:
            sheet = book[sheet_name]

        # add data row
        product_id = response.meta.get('product_id')
        row = [product_id, itemlist['title'], itemlist['image'], itemlist['catalog_number'], itemlist['Check Stock'], itemlist['desc']]
        sheet.append(row)

        # Save the workbook
        book.save('datalist.xlsx')
